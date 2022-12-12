import openpyxl
from glob import glob
import argparse
import os

# Set argparse
parser = argparse.ArgumentParser(description='Sum_excel')


parser.add_argument('--data_dir', metavar='DIR', default='./output/V4_ec/90_mount',
                    help='path to excels')


def create_matrix1(data):
    sheet = data.worksheets[0]
    table = list(sheet.values)
    res = []
    for v in table[0]:
        res.append(v)
    """
        Single validation dataset
    """
    # For single excel records 9 values
    # matrix1.append([res[0], res[6], res[7] * 100, res[8] * 100])
    # For single excel records 3 values
    matrix1.append([res[0], res[1], res[2] * 100, res[3] * 100])

    """
        Two or more validation dataset
    """
    # matrix1.append([res[0], res[1], res[2] * 100, res[3] * 100, res[4], res[5], res[6] * 100, res[7] * 100])
    # matrix1.append([res[0], res[1], res[2] * 100, res[3] * 100, res[4], res[5], res[6] * 100, res[7] * 100,res[8], res[9], res[10] * 100, res[11] * 100])


    return res[0]


def create_matrix2(data, model_name):
    sheet = data.worksheets[2]
    table = list(sheet.values)

    if len(matrix2) == 0:
        matrix2.append(['Ground Truth', model_name])
        for line in table:
            a = list(line)
            matrix2.append(a)
    else:
        matrix2[0].append(model_name)
        for i, line in enumerate(table):
            a = list(line)
            matrix2[i+1].append(a[1])


def create_matrix3(data, model_name):
    sheet = data.worksheets[3]
    table = list(sheet.values)

    if len(matrix3) == 0:
        matrix3.append([model_name+'_train_loss', 'test_loss'])
        for line in table:
            a = list(line)
            matrix3.append(a)
    else:
        matrix3[0].append(model_name+'_train_loss')
        matrix3[0].append('test_loss')
        for i, line in enumerate(table):
            a = list(line)
            matrix3[i+1].append(a[0])
            matrix3[i + 1].append(a[1])


if __name__ == "__main__":
    # get all args params
    args = parser.parse_args()
    out_file = os.path.join(args.data_dir, 'result.xlsx')

    # Global var
    matrix1 = []
    matrix2 = []
    matrix3 = []
    for root, dirs_name, files_name in os.walk(args.data_dir):
        for i in files_name:
            if i.split('.')[-1] in ['xlsx']:
                file_name = os.path.join(root, i)
                data = openpyxl.load_workbook(file_name)

                # for sheet 1
                model_name = create_matrix1(data)

                # for sheet 2
                create_matrix2(data, model_name)

                # for sheet 2
                create_matrix3(data, model_name)


    # matrix_t = list(map(list,zip(*matrix)))

    wb = openpyxl.Workbook()
    s1 = wb.get_sheet_by_name('Sheet')
    s2 = wb.create_sheet('Sheet2')
    s3 = wb.create_sheet('Sheet3')
    for line in matrix1:
        s1.append(line)
    for line in matrix2:
        s2.append(line)
    for line in matrix3:
        s3.append(line)
    wb.save(out_file)